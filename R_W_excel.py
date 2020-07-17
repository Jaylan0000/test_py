# -*- coding:utf-8 -*-
# @time   :2020/6/30  15:19
# @Author :feiou
# @Email  :1006290526@qq.com
# @File   :R_W_excel.py

#用来读取测试数据的

#R 代表read，W代表write   读取所有的数据
#在桌面保存的Excel文件要复制到python里面来
from openpyxl import load_workbook  #打开工作簿

def read_data(file_name,sheet_name):
    wb = load_workbook(file_name)  #打开工作簿
    sheet = wb[sheet_name]  #定位表单
    all_case=[]  #存储所有的数据
    for i in range(1,sheet.max_row): #获取所有的行
        case=[]  #存储每一行数据
        for j in range(1,sheet.max_column-1):  #存储所有的列
            case.append(sheet.cell(row=i+1,column=j).value)  #将获取到的每一行的数据添加到case[]列表中
            # print(sheet.cell(row=i+1,column=j).value)
        all_case.append(case)  #将获取到的所有的数据按行读取  添加到all_case列表中
    return all_case  #注意缩进  返回all_case给run.py用

def write_data(file_name,sheet_name,row,column,value):#此函数是写入到Excel中
    # 开始写入结果
    wb = load_workbook(file_name)
    sheet = wb[sheet_name]
    # 定位单元格存值  行 列 值  除了跟着i走，还可以跟着id走
    sheet.cell(row=row, column=column).value = value  # 前面的http_request函数里是有返回结果的

    wb.save(file_name)  # 要保存一下

if __name__ == '__main__':  # 只有你在当前文件夹下面才可以执行这个数据  避免在其他地方调用的时候打印两次
    all_case_1=read_data('test_data.xlsx','recharge')
    print(all_case_1)  #打印所有的数据
